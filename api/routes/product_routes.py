from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity, verify_jwt_in_request
from database import db
from models.product import Product
from models.shop import Shop
from models.user import User
import os

product_bp = Blueprint('products', __name__)

def get_optional_user():
    try:
        verify_jwt_in_request(optional=True)
        uid = get_jwt_identity()
        if uid:
            return User.query.get(int(uid))
    except Exception:
        pass
    return None

# ─── Get All Products ─────────────────────────────────────────────────────────
@product_bp.route('', methods=['GET'])
def get_products():
    search = request.args.get('search', '')
    category = request.args.get('category', '')
    shop_id = request.args.get('shopId', '')
    min_price = request.args.get('minPrice', '')
    max_price = request.args.get('maxPrice', '')
    sort_by = request.args.get('sortBy', '')

    query = Product.query.join(Shop).filter(
        Product.is_active == True,
        Shop.is_active == True,
        Shop.is_approved == True
    )

    if search:
        like = f'%{search}%'
        query = query.filter(
            db.or_(
                Product.name.ilike(like),
                Product.description.ilike(like),
                Product.category.ilike(like),
                Product.subcategory.ilike(like)
            )
        )

    if category and category != 'All':
        query = query.filter(Product.category == category)

    if shop_id:
        query = query.filter(Product.shop_id == int(shop_id))

    if min_price:
        query = query.filter(Product.price >= float(min_price))
    if max_price:
        query = query.filter(Product.price <= float(max_price))

    if sort_by == 'price_asc':
        query = query.order_by(Product.price.asc())
    elif sort_by == 'price_desc':
        query = query.order_by(Product.price.desc())
    elif sort_by == 'views':
        query = query.order_by(Product.views.desc())
    else:
        query = query.order_by(Product.created_at.desc())

    products = query.all()
    return jsonify([p.to_dict() for p in products]), 200

# ─── Compare Products ─────────────────────────────────────────────────────────
@product_bp.route('/compare', methods=['GET'])
def compare_products():
    category = request.args.get('category')
    search = request.args.get('search', '')
    min_price = request.args.get('minPrice', '')
    max_price = request.args.get('maxPrice', '')

    if not category:
        return jsonify({'message': 'Category is required for comparison'}), 400

    query = Product.query.filter_by(is_active=True, category=category)
    if search:
        query = query.filter(Product.name.ilike(f'%{search}%'))
    if min_price:
        query = query.filter(Product.price >= float(min_price))
    if max_price:
        query = query.filter(Product.price <= float(max_price))

    products = query.order_by(Product.price.asc()).all()
    return jsonify([p.to_dict() for p in products]), 200

# ─── Get Product By ID ────────────────────────────────────────────────────────
@product_bp.route('/<int:product_id>', methods=['GET'])
def get_product_by_id(product_id):
    product = Product.query.get(product_id)
    if not product:
        return jsonify({'message': 'Product not found'}), 404
    product.views += 1
    db.session.commit()
    return jsonify(product.to_dict()), 200

# ─── Create Product ───────────────────────────────────────────────────────────
@product_bp.route('', methods=['POST'])
@jwt_required()
def create_product():
    user_id = int(get_jwt_identity())
    data = request.get_json()

    shop = Shop.query.filter_by(owner_id=user_id).first()
    if not shop:
        return jsonify({'message': 'No shop found for this user'}), 404

    product = Product(
        name=data.get('name'),
        description=data.get('description'),
        price=float(data.get('price', 0)),
        offer_price=float(data['offerPrice']) if data.get('offerPrice') else None,
        stock=int(data.get('stock', 0)),
        image_url=data.get('imageUrl'),
        category=data.get('category'),
        subcategory=data.get('subcategory'),
        shop_id=shop.id
    )
    db.session.add(product)
    db.session.commit()
    return jsonify(product.to_dict()), 201

# ─── Update Product ───────────────────────────────────────────────────────────
@product_bp.route('/<int:product_id>', methods=['PUT'])
@jwt_required()
def update_product(product_id):
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    product = Product.query.get(product_id)

    if not product:
        return jsonify({'message': 'Product not found'}), 404
    if product.shop.owner_id != user_id and user.role != 'ADMIN':
        return jsonify({'message': 'Not authorized'}), 401

    data = request.get_json()
    if 'name' in data: product.name = data['name']
    if 'description' in data: product.description = data['description']
    if 'price' in data: product.price = float(data['price'])
    if 'offerPrice' in data: product.offer_price = float(data['offerPrice']) if data['offerPrice'] else None
    if 'stock' in data: product.stock = int(data['stock'])
    if 'imageUrl' in data: product.image_url = data['imageUrl']
    if 'category' in data: product.category = data['category']
    if 'subcategory' in data: product.subcategory = data['subcategory']
    if 'isActive' in data: product.is_active = bool(data['isActive'])

    db.session.commit()
    return jsonify(product.to_dict()), 200

# ─── Delete Product ───────────────────────────────────────────────────────────
@product_bp.route('/<int:product_id>', methods=['DELETE'])
@jwt_required()
def delete_product(product_id):
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    product = Product.query.get(product_id)

    if not product:
        return jsonify({'message': 'Product not found'}), 404
    if product.shop.owner_id != user_id and user.role != 'ADMIN':
        return jsonify({'message': 'Not authorized'}), 401

    db.session.delete(product)
    db.session.commit()
    return jsonify({'message': 'Product removed'}), 200

# ─── Bulk Upload Products ─────────────────────────────────────────────────────
@product_bp.route('/bulk-upload', methods=['POST'])
@jwt_required()
def bulk_upload_products():
    import io
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        user_id = int(get_jwt_identity())
        logger.info(f"Bulk upload started by user {user_id}")
        
        shop = Shop.query.filter_by(owner_id=user_id).first()
        if not shop:
            logger.warning(f"User {user_id} has no shop")
            return jsonify({'message': 'You must create a shop first before uploading products. Go to "Register Your Shop" and setup your shop.'}), 404

        if 'file' not in request.files:
            logger.warning(f"No file in request from user {user_id}")
            return jsonify({'message': 'Please upload an Excel (.xlsx, .xls) or CSV (.csv) file'}), 400

        file = request.files['file']
        
        if not file or file.filename == '':
            logger.warning(f"Empty filename from user {user_id}")
            return jsonify({'message': 'No file selected'}), 400
        
        # Validate file extension
        valid_extensions = {'.xlsx', '.xls', '.csv'}
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in valid_extensions:
            logger.warning(f"Invalid file type: {file_ext} from user {user_id}")
            return jsonify({'message': f'Invalid file type. Please upload .xlsx, .xls, or .csv file'}), 400
        
        try:
            import openpyxl
            logger.info(f"Parsing file: {file.filename}")
            file_content = file.read()
            file.seek(0)  # Reset file pointer
            
            wb = openpyxl.load_workbook(io.BytesIO(file_content))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            
            # Normalize header names (remove spaces, lowercase)
            normalized_headers = {}
            for header in headers:
                if header:
                    normalized = header.strip().lower().replace(' ', '')
                    normalized_headers[normalized] = header
            
            logger.info(f"Headers found: {headers}")
            logger.info(f"Normalized headers: {list(normalized_headers.keys())}")
            
            # Check for required columns
            if 'name' not in normalized_headers and 'productname' not in normalized_headers and 'itemname' not in normalized_headers:
                return jsonify({'message': 'Missing required "name" column in your spreadsheet'}), 400
            if 'price' not in normalized_headers:
                return jsonify({'message': 'Missing required "price" column in your spreadsheet'}), 400
            
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            
            logger.info(f"File parsed successfully. Found {len(rows)} rows")
        except Exception as e:
            logger.error(f"Failed to parse file: {str(e)}")
            return jsonify({'message': f'Failed to parse file: {str(e)}'}), 400

        created = []
        errors = []
        
        # Helper function to get value from row using flexible column names
        def get_row_value(row, *possible_keys):
            """Get value from row using multiple possible column names"""
            for key in possible_keys:
                if key in row and row[key] is not None:
                    return row[key]
            return None
        
        for idx, row in enumerate(rows, start=2):
            try:
                # Get values with flexible column name matching
                name = get_row_value(row, 'name', 'Name', 'product name', 'productname', 'Product Name', 'item name', 'itemname')
                price = get_row_value(row, 'price', 'Price', 'unit price', 'unitprice', 'Unit Price')
                description = get_row_value(row, 'description', 'Description', 'desc', 'Desc')
                offer_price = get_row_value(row, 'offerPrice', 'offer price', 'offerprice', 'Offer Price', 'discount price', 'discountprice')
                stock = get_row_value(row, 'stock', 'Stock', 'quantity', 'Quantity', 'qty', 'Qty')
                image_url = get_row_value(row, 'imageUrl', 'image url', 'imageurl', 'Image URL', 'image', 'Image')
                category = get_row_value(row, 'category', 'Category', 'cat', 'Cat')
                subcategory = get_row_value(row, 'subcategory', 'sub category', 'subcategory', 'Subcategory', 'Sub Category')
                
                if not name or not price:
                    errors.append({'row': idx, 'error': 'Missing required fields (name, price)'})
                    continue
                
                try:
                    price_float = float(price)
                except (ValueError, TypeError):
                    errors.append({'row': idx, 'error': f'Invalid price: {price} (must be a number)'})
                    continue
                
                p = Product(
                    name=str(name).strip(),
                    description=str(description).strip() if description else '',
                    price=price_float,
                    offer_price=float(offer_price) if offer_price else None,
                    stock=int(float(stock)) if stock else 0,
                    image_url=str(image_url).strip() if image_url else '',
                    category=str(category).strip() if category else 'General',
                    subcategory=str(subcategory).strip() if subcategory else '',
                    shop_id=shop.id
                )
                db.session.add(p)
                created.append(p)
            except Exception as e:
                logger.error(f"Error processing row {idx}: {str(e)}")
                errors.append({'row': idx, 'error': str(e)})

        db.session.commit()
        logger.info(f"Bulk upload complete: {len(created)} created, {len(errors)} errors")
        
        return jsonify({
            'message': f'Processed {len(rows)} items',
            'successCount': len(created),
            'failedCount': len(errors),
            'errors': errors
        }), 200
    except Exception as e:
        logger.error(f"Bulk upload failed: {str(e)}", exc_info=True)
        return jsonify({'message': f'Server error: {str(e)}'}), 500
