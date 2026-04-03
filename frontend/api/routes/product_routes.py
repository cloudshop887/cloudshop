from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity, verify_jwt_in_request
from database import db
from models.product import Product
from models.shop import Shop
from models.user import User
import io
import csv

product_bp = Blueprint('products', __name__)

def get_optional_user():
    try:
        verify_jwt_in_request(optional=True)
        uid = get_jwt_identity()
        if uid:
            return User.query.get(int(uid))
    except Exception:
        pass
    return None

# ─── Get All Products ─────────────────────────────────────────────────────────
@product_bp.route('', methods=['GET'])
def get_products():
    search = request.args.get('search', '')
    category = request.args.get('category', '')
    shop_id = request.args.get('shopId', '')
    min_price = request.args.get('minPrice', '')
    max_price = request.args.get('maxPrice', '')
    sort_by = request.args.get('sortBy', '')

    query = Product.query.join(Shop).filter(
        Product.is_active == True,
        Shop.is_active == True,
        Shop.is_approved == True
    )

    if search:
        like = f'%{search}%'
        query = query.filter(
            db.or_(
                Product.name.ilike(like),
                Product.description.ilike(like),
                Product.category.ilike(like),
                Product.subcategory.ilike(like)
            )
        )

    if category and category != 'All':
        query = query.filter(Product.category == category)

    if shop_id:
        query = query.filter(Product.shop_id == int(shop_id))

    if min_price:
        query = query.filter(Product.price >= float(min_price))
    if max_price:
        query = query.filter(Product.price <= float(max_price))

    if sort_by == 'price_asc':
        query = query.order_by(Product.price.asc())
    elif sort_by == 'price_desc':
        query = query.order_by(Product.price.desc())
    elif sort_by == 'views':
        query = query.order_by(Product.views.desc())
    else:
        query = query.order_by(Product.created_at.desc())

    products = query.all()
    return jsonify([p.to_dict() for p in products]), 200

# ─── Compare Products ─────────────────────────────────────────────────────────
@product_bp.route('/compare', methods=['GET'])
def compare_products():
    category = request.args.get('category')
    search = request.args.get('search', '')
    min_price = request.args.get('minPrice', '')
    max_price = request.args.get('maxPrice', '')

    if not category:
        return jsonify({'message': 'Category is required for comparison'}), 400

    query = Product.query.filter_by(is_active=True, category=category)
    if search:
        query = query.filter(Product.name.ilike(f'%{search}%'))
    if min_price:
        query = query.filter(Product.price >= float(min_price))
    if max_price:
        query = query.filter(Product.price <= float(max_price))

    products = query.order_by(Product.price.asc()).all()
    return jsonify([p.to_dict() for p in products]), 200

# ─── Get Product By ID ────────────────────────────────────────────────────────
@product_bp.route('/<int:product_id>', methods=['GET'])
def get_product_by_id(product_id):
    product = Product.query.get(product_id)
    if not product:
        return jsonify({'message': 'Product not found'}), 404
    product.views += 1
    db.session.commit()
    return jsonify(product.to_dict()), 200

# ─── Create Product ───────────────────────────────────────────────────────────
@product_bp.route('', methods=['POST'])
@jwt_required()
def create_product():
    user_id = int(get_jwt_identity())
    data = request.get_json()

    shop = Shop.query.filter_by(owner_id=user_id).first()
    if not shop:
        return jsonify({'message': 'No shop found for this user'}), 404

    product = Product(
        name=data.get('name'),
        description=data.get('description'),
        price=float(data.get('price', 0)),
        offer_price=float(data['offerPrice']) if data.get('offerPrice') else None,
        stock=int(data.get('stock', 0)),
        image_url=data.get('imageUrl'),
        category=data.get('category'),
        subcategory=data.get('subcategory'),
        shop_id=shop.id
    )
    db.session.add(product)
    db.session.commit()
    return jsonify(product.to_dict()), 201

# ─── Update Product ───────────────────────────────────────────────────────────
@product_bp.route('/<int:product_id>', methods=['PUT'])
@jwt_required()
def update_product(product_id):
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    product = Product.query.get(product_id)

    if not product:
        return jsonify({'message': 'Product not found'}), 404
    if product.shop.owner_id != user_id and user.role != 'ADMIN':
        return jsonify({'message': 'Not authorized'}), 401

    data = request.get_json()
    if 'name' in data: product.name = data['name']
    if 'description' in data: product.description = data['description']
    if 'price' in data: product.price = float(data['price'])
    if 'offerPrice' in data: product.offer_price = float(data['offerPrice']) if data['offerPrice'] else None
    if 'stock' in data: product.stock = int(data['stock'])
    if 'imageUrl' in data: product.image_url = data['imageUrl']
    if 'category' in data: product.category = data['category']
    if 'subcategory' in data: product.subcategory = data['subcategory']
    if 'isActive' in data: product.is_active = bool(data['isActive'])

    db.session.commit()
    return jsonify(product.to_dict()), 200

# ─── Delete Product ───────────────────────────────────────────────────────────
@product_bp.route('/<int:product_id>', methods=['DELETE'])
@jwt_required()
def delete_product(product_id):
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    product = Product.query.get(product_id)

    if not product:
        return jsonify({'message': 'Product not found'}), 404
    if product.shop.owner_id != user_id and user.role != 'ADMIN':
        return jsonify({'message': 'Not authorized'}), 401

    db.session.delete(product)
    db.session.commit()
    return jsonify({'message': 'Product removed'}), 200

# ─── Bulk Upload Products ─────────────────────────────────────────────────────
@product_bp.route('/bulk-upload', methods=['POST'])
@jwt_required()
def bulk_upload_products():
    user_id = int(get_jwt_identity())
    shop = Shop.query.filter_by(owner_id=user_id).first()
    if not shop:
        return jsonify({'message': 'No shop found for this user'}), 404

    if 'file' not in request.files:
        return jsonify({'message': 'No file uploaded. Please attach a file with field name "file"'}), 400

    file = request.files['file']
    filename = file.filename.lower() if file.filename else ''

    # ── Detect file type by extension (more reliable than MIME type) ──
    rows = []
    if filename.endswith('.csv'):
        # ✅ Handle CSV
        try:
            content = file.read().decode('utf-8-sig')  # utf-8-sig strips BOM if present
            reader = csv.DictReader(io.StringIO(content))
            rows = [row for row in reader]
        except UnicodeDecodeError:
            try:
                file.seek(0)
                content = file.read().decode('latin-1')
                reader = csv.DictReader(io.StringIO(content))
                rows = [row for row in reader]
            except Exception as e:
                return jsonify({'message': f'Failed to read CSV file: {str(e)}'}), 400
        except Exception as e:
            return jsonify({'message': f'Failed to parse CSV: {str(e)}'}), 400

    elif filename.endswith('.xlsx') or filename.endswith('.xls'):
        # ✅ Handle Excel
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file.read()))
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value is not None else '' for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Skip completely empty rows
                if any(v is not None for v in row):
                    rows.append(dict(zip(headers, row)))
        except Exception as e:
            return jsonify({'message': f'Failed to parse Excel file: {str(e)}'}), 400

    else:
        # ✅ Clear error message telling user what's wrong
        return jsonify({
            'message': f'Unsupported file type "{file.filename}". Please upload a .csv, .xlsx, or .xls file.'
        }), 400

    if not rows:
        return jsonify({'message': 'File is empty or has no data rows'}), 400

    # ── Process rows ──────────────────────────────────────────────────
    created = []
    errors = []

    for i, row in enumerate(rows, start=2):  # start=2 because row 1 is header
        try:
            name = str(row.get('name', '') or '').strip()
            price_raw = row.get('price', '')

            if not name:
                errors.append({'row': i, 'error': 'Missing required field: name'})
                continue
            if not price_raw:
                errors.append({'row': i, 'error': 'Missing required field: price'})
                continue

            p = Product(
                name=name,
                description=str(row.get('description', '') or '').strip(),
                price=float(price_raw),
                offer_price=float(row['offerPrice']) if row.get('offerPrice') else None,
                stock=int(row['stock']) if row.get('stock') else 0,
                image_url=str(row.get('imageUrl', '') or '').strip(),
                category=str(row.get('category', 'General') or 'General').strip(),
                subcategory=str(row.get('subcategory', '') or '').strip(),
                shop_id=shop.id
            )
            db.session.add(p)
            created.append(p)

        except Exception as e:
            errors.append({'row': i, 'error': str(e)})

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Database error: {str(e)}'}), 500

    return jsonify({
        'message': f'Processed {len(rows)} rows',
        'successCount': len(created),
        'failedCount': len(errors),
        'errors': errors
    }), 200