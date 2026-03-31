from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity, verify_jwt_in_request
from database import db
from models.product import Product
from models.shop import Shop
from models.user import User

product_bp = Blueprint('products', __name__)

def get_optional_user():
    try:
        verify_jwt_in_request(optional=True)
        uid = get_jwt_identity()
        if uid:
            return User.query.get(int(uid))
    except Exception:
        pass
    return None

# ─── Get All Products ─────────────────────────────────────────────────────────
@product_bp.route('', methods=['GET'])
def get_products():
    search = request.args.get('search', '')
    category = request.args.get('category', '')
    shop_id = request.args.get('shopId', '')
    min_price = request.args.get('minPrice', '')
    max_price = request.args.get('maxPrice', '')
    sort_by = request.args.get('sortBy', '')

    query = Product.query.join(Shop).filter(
        Product.is_active == True,
        Shop.is_active == True,
        Shop.is_approved == True
    )

    if search:
        like = f'%{search}%'
        query = query.filter(
            db.or_(
                Product.name.ilike(like),
                Product.description.ilike(like),
                Product.category.ilike(like),
                Product.subcategory.ilike(like)
            )
        )

    if category and category != 'All':
        query = query.filter(Product.category == category)

    if shop_id:
        query = query.filter(Product.shop_id == int(shop_id))

    if min_price:
        query = query.filter(Product.price >= float(min_price))
    if max_price:
        query = query.filter(Product.price <= float(max_price))

    if sort_by == 'price_asc':
        query = query.order_by(Product.price.asc())
    elif sort_by == 'price_desc':
        query = query.order_by(Product.price.desc())
    elif sort_by == 'views':
        query = query.order_by(Product.views.desc())
    else:
        query = query.order_by(Product.created_at.desc())

    products = query.all()
    return jsonify([p.to_dict() for p in products]), 200

# ─── Compare Products ─────────────────────────────────────────────────────────
@product_bp.route('/compare', methods=['GET'])
def compare_products():
    category = request.args.get('category')
    search = request.args.get('search', '')
    min_price = request.args.get('minPrice', '')
    max_price = request.args.get('maxPrice', '')

    if not category:
        return jsonify({'message': 'Category is required for comparison'}), 400

    query = Product.query.filter_by(is_active=True, category=category)
    if search:
        query = query.filter(Product.name.ilike(f'%{search}%'))
    if min_price:
        query = query.filter(Product.price >= float(min_price))
    if max_price:
        query = query.filter(Product.price <= float(max_price))

    products = query.order_by(Product.price.asc()).all()
    return jsonify([p.to_dict() for p in products]), 200

# ─── Get Product By ID ────────────────────────────────────────────────────────
@product_bp.route('/<int:product_id>', methods=['GET'])
def get_product_by_id(product_id):
    product = Product.query.get(product_id)
    if not product:
        return jsonify({'message': 'Product not found'}), 404
    product.views += 1
    db.session.commit()
    return jsonify(product.to_dict()), 200

# ─── Create Product ───────────────────────────────────────────────────────────
@product_bp.route('', methods=['POST'])
@jwt_required()
def create_product():
    user_id = int(get_jwt_identity())
    data = request.get_json()

    shop = Shop.query.filter_by(owner_id=user_id).first()
    if not shop:
        return jsonify({'message': 'No shop found for this user'}), 404

    product = Product(
        name=data.get('name'),
        description=data.get('description'),
        price=float(data.get('price', 0)),
        offer_price=float(data['offerPrice']) if data.get('offerPrice') else None,
        stock=int(data.get('stock', 0)),
        image_url=data.get('imageUrl'),
        category=data.get('category'),
        subcategory=data.get('subcategory'),
        shop_id=shop.id
    )
    db.session.add(product)
    db.session.commit()
    return jsonify(product.to_dict()), 201

# ─── Update Product ───────────────────────────────────────────────────────────
@product_bp.route('/<int:product_id>', methods=['PUT'])
@jwt_required()
def update_product(product_id):
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    product = Product.query.get(product_id)

    if not product:
        return jsonify({'message': 'Product not found'}), 404
    if product.shop.owner_id != user_id and user.role != 'ADMIN':
        return jsonify({'message': 'Not authorized'}), 401

    data = request.get_json()
    if 'name' in data: product.name = data['name']
    if 'description' in data: product.description = data['description']
    if 'price' in data: product.price = float(data['price'])
    if 'offerPrice' in data: product.offer_price = float(data['offerPrice']) if data['offerPrice'] else None
    if 'stock' in data: product.stock = int(data['stock'])
    if 'imageUrl' in data: product.image_url = data['imageUrl']
    if 'category' in data: product.category = data['category']
    if 'subcategory' in data: product.subcategory = data['subcategory']
    if 'isActive' in data: product.is_active = bool(data['isActive'])

    db.session.commit()
    return jsonify(product.to_dict()), 200

# ─── Delete Product ───────────────────────────────────────────────────────────
@product_bp.route('/<int:product_id>', methods=['DELETE'])
@jwt_required()
def delete_product(product_id):
    user_id = int(get_jwt_identity())
    user = User.query.get(user_id)
    product = Product.query.get(product_id)

    if not product:
        return jsonify({'message': 'Product not found'}), 404
    if product.shop.owner_id != user_id and user.role != 'ADMIN':
        return jsonify({'message': 'Not authorized'}), 401

    db.session.delete(product)
    db.session.commit()
    return jsonify({'message': 'Product removed'}), 200

# ─── Bulk Upload Products ─────────────────────────────────────────────────────
@product_bp.route('/bulk-upload', methods=['POST'])
@jwt_required()
def bulk_upload_products():
    user_id = int(get_jwt_identity())
    shop = Shop.query.filter_by(owner_id=user_id).first()
    if not shop:
        return jsonify({'message': 'No shop found for this user'}), 404

    if 'file' not in request.files:
        return jsonify({'message': 'Please upload an Excel (.xlsx, .xls) or CSV (.csv) file'}), 400

    file = request.files['file']
    import io
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
    except Exception:
        return jsonify({'message': 'Failed to parse file'}), 400

    created = []
    errors = []
    for row in rows:
        try:
            if not row.get('name') or not row.get('price'):
                errors.append({'row': row, 'error': 'Missing required fields'})
                continue
            p = Product(
                name=row['name'],
                description=row.get('description', ''),
                price=float(row['price']),
                offer_price=float(row['offerPrice']) if row.get('offerPrice') else None,
                stock=int(row['stock']) if row.get('stock') else 0,
                image_url=row.get('imageUrl', ''),
                category=row.get('category', 'General'),
                subcategory=row.get('subcategory', ''),
                shop_id=shop.id
            )
            db.session.add(p)
            created.append(p)
        except Exception as e:
            errors.append({'row': row, 'error': str(e)})

    db.session.commit()
    return jsonify({
        'message': f'Processed {len(rows)} items',
        'successCount': len(created),
        'failedCount': len(errors),
        'errors': errors
    }), 200
